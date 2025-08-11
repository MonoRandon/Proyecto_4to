import cv2
import numpy as np
from tkinter import messagebox
from PIL import ImageGrab
import os
import openpyxl
import pytesseract
import math
from collections import defaultdict


start_x = None
start_y = None

def angle_between(p1, p2):
    """Calcula el ángulo en grados entre dos puntos p1 y p2."""
    x1, y1 = p1
    x2, y2 = p2
    ang_rad = math.atan2(y2 - y1, x2 - x1)
    ang_deg = math.degrees(ang_rad)
    return ang_deg

class DrawingHelper:
    def analizar_dibujo(self, canvas, area_texto):
        """Analiza el dibujo del plano usando OpenCV y lo guarda corregido con sugerencias."""
        try:
            img = cv2.imread("dibujo.png")
            gris = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            bordes = cv2.Canny(gris, 50, 150)
            lineas = cv2.HoughLinesP(bordes, 1, np.pi / 180, 100, minLineLength=50, maxLineGap=5)

            if lineas is not None:
                for linea in lineas:
                    x1, y1, x2, y2 = linea[0]
                    cv2.line(img, (x1, y1), (x2, y2), (0, 255, 0), 2)

            cv2.imwrite("plano_corregido.png", img)

            sugerencias = []

            if lineas is not None:
                num_lineas = len(lineas)
                sugerencias.append(f"Se detectaron {num_lineas} líneas.")

                if num_lineas > 20:
                    sugerencias.append("Podrías simplificar el diseño eliminando líneas innecesarias.")
                elif num_lineas < 5:
                    sugerencias.append("El plano parece incompleto. ¿Deseas añadir más detalles?")
                else:
                    sugerencias.append("El número de líneas parece adecuado para un plano básico.")
            else:
                sugerencias.append("No se detectaron líneas. ¿Estás segur@ de que dibujaste algo?")

            mensaje_final = "Asistente: Análisis completo.\n" + "\n".join(sugerencias) + "\nEsperando órdenes para complementarlo.\n"
            area_texto.insert("end", mensaje_final)

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo procesar el plano: {e}")

def start_draw(event):
    global start_x, start_y
    start_x = event.x
    start_y = event.y

def draw(event, canvas):
    global start_x, start_y
    canvas.create_line(start_x, start_y, event.x, event.y, fill="black", width=2)
    start_x = event.x
    start_y = event.y

def activar_dibujo(canvas, text_output, root, drawing_helper_instance):
    """Permite dibujar con el mouse y al soltar analiza automáticamente."""
    canvas.bind("<Button-1>", start_draw)
    canvas.bind("<B1-Motion>", lambda e: draw(e, canvas))
    canvas.bind("<ButtonRelease-1>", lambda e: guardar_y_analizar(canvas, text_output, root, drawing_helper_instance))

def guardar_y_analizar(canvas, text_output, root, drawing_helper_instance):
    """Guarda el canvas como imagen y llama al análisis de dibujo."""
    x = root.winfo_rootx() + canvas.winfo_x()
    y = root.winfo_rooty() + canvas.winfo_y()
    x1 = x + canvas.winfo_width()
    y1 = y + canvas.winfo_height()

    img = ImageGrab.grab().crop((x, y, x1, y1))
    img.save("dibujo.png")

    drawing_helper_instance.analizar_dibujo(canvas, text_output)

def guardar_canvas(canvas, root, nombre="dibujo.png"):
    x = root.winfo_rootx() + canvas.winfo_x()
    y = root.winfo_rooty() + canvas.winfo_y()
    x1 = x + canvas.winfo_width()
    y1 = y + canvas.winfo_height()
    img = ImageGrab.grab().crop((x, y, x1, y1))
    img.save(nombre)

def analizar_dibujo(canvas, text_output, root):
    guardar_canvas(canvas, root)
    text_output.insert("end", "Asistente: Dibujo analizado. Esperando tu orden para continuar...\n")

def activar_dibujo(canvas, text_output, root):
    canvas.bind("<Button-1>", start_draw)
    canvas.bind("<B1-Motion>", lambda e: draw(e, canvas))
    canvas.bind("<ButtonRelease-1>", lambda e: analizar_dibujo(canvas, text_output, root))

def dibujar_gantt(canvas, tareas):
    canvas.delete("all")  # Limpiar antes

    x_inicio = 100
    y_inicio = 50
    ancho_dia = 30
    alto_fila = 40

    # Dibujar encabezados (días)
    for i in range(15):
        x = x_inicio + i * ancho_dia
        canvas.create_text(x + ancho_dia / 2, y_inicio - 20, text=f"Día {i+1}", font=("Arial", 10))

    # Dibujar tareas
    for idx, (tarea, inicio, duracion) in enumerate(tareas):
        y = y_inicio + idx * alto_fila
        canvas.create_text(x_inicio - 70, y + alto_fila / 2, text=tarea, anchor="w", font=("Arial", 10, "bold"))

        for d in range(duracion):
            x = x_inicio + (inicio + d) * ancho_dia
            canvas.create_rectangle(x, y, x + ancho_dia, y + alto_fila - 10, fill="#4F81BD", outline="black")

def analizar_plano(ruta_imagen):
    try:
        imagen = cv2.imread(ruta_imagen)
        gris = cv2.cvtColor(imagen, cv2.COLOR_BGR2GRAY)
        borroso = cv2.GaussianBlur(gris, (5, 5), 0)
        bordes = cv2.Canny(borroso, 50, 150)

        contornos, _ = cv2.findContours(bordes, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        rectangulos = 0
        lineas_horizontales = 0
        lineas_verticales = 0

        for c in contornos:
            aprox = cv2.approxPolyDP(c, 0.02 * cv2.arcLength(c, True), True)
            x, y, w, h = cv2.boundingRect(aprox)

            if len(aprox) == 4:
                rectangulos += 1
                aspect_ratio = w / float(h)
                if aspect_ratio > 5:  # muy ancho → línea horizontal
                    lineas_horizontales += 1
                elif aspect_ratio < 0.2:  # muy alto → línea vertical
                    lineas_verticales += 1

        texto_analisis = (
            f"Se detectaron {rectangulos} figuras rectangulares.\n"
            f"Entre ellas, hay {lineas_horizontales} líneas horizontales "
            f"y {lineas_verticales} líneas verticales.\n"
        )

        if rectangulos == 0:
            texto_analisis += "Podría ser un dibujo libre o incompleto.\n"
        elif rectangulos >= 5:
            texto_analisis += "Parece una estructura con varias secciones. Quizá una tabla o presentación.\n"
        elif lineas_verticales >= 2:
            texto_analisis += "Hay divisiones verticales, puede que sea una columna de tareas.\n"
        else:
            texto_analisis += "El contenido es básico, puede tratarse de un esquema o bosquejo."

        return texto_analisis

    except Exception as e:
        return f"Error al analizar la imagen: {e}"

def leer_tareas_excel(carpeta, nombre_archivo):
    ruta = os.path.join(carpeta, nombre_archivo)
    if not os.path.exists(ruta):
        return "Archivo no encontrado."

    wb = openpyxl.load_workbook(ruta)
    ws = wb.active

    tareas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[0]:  # Solo si hay nombre de tarea
            tareas.append(f"Tarea: {fila[0]}, Inicio: {fila[1]}, Fin: {fila[2]}")
    return "\n".join(tareas) if tareas else "No hay tareas en el archivo."

def detectar_texto(imagen):
    return pytesseract.image_to_string(imagen)

def obtener_tareas_para_gantt(carpeta, nombre_archivo):
    ruta = os.path.join(carpeta, nombre_archivo)
    if not os.path.exists(ruta):
        return "Archivo no encontrado."

    wb = openpyxl.load_workbook(ruta)
    ws = wb.active

    tareas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[0] and fila[1] and fila[2]:
            tareas.append({
                "nombre": fila[0],
                "inicio": fila[1],
                "fin": fila[2]
            })
    return tareas if tareas else "No hay tareas en el archivo."

def analizar_y_corregir(ruta_entrada, salida_dir="static/uploads", basename="proyecto"):
    """
    Analiza la imagen de entrada con OpenCV y genera:
      - imagen corregida (se marca en rojo errores y en verde correcciones)
      - resumen técnico con conteos, desviaciones, ángulos problemáticos, etc.
    Retorna: ruta_corregida, resumen (dict)
    """
    img = cv2.imread(ruta_entrada)
    if img is None:
        raise FileNotFoundError(f"No se encontró {ruta_entrada}")

    h, w = img.shape[:2]
    gris = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # mejora contraste si hace falta
    gris = cv2.equalizeHist(gris)

    # Canny para bordes
    bordes = cv2.Canny(gris, 50, 150, apertureSize=3)

    # Hough probabilista para líneas
    lines = cv2.HoughLinesP(bordes, rho=1, theta=np.pi/180, threshold=60, minLineLength=40, maxLineGap=8)

    resumen = {
        "num_lineas": 0,
        "lineas": [],  # lista de dicts {x1,y1,x2,y2, angle}
        "problemas": [],
        "estadisticas": {}
    }

    if lines is None:
        resumen["num_lineas"] = 0
        resumen["problemas"].append("No se detectaron líneas claras. Revisa el trazo o contraste.")
    else:
        resumen["num_lineas"] = len(lines)
        # calcular ángulos y bounding boxes
        angles = []
        for l in lines:
            x1,y1,x2,y2 = l[0]
            ang = angle_between((x1,y1),(x2,y2))
            angles.append(ang)
            resumen["lineas"].append({"x1":int(x1),"y1":int(y1),"x2":int(x2),"y2":int(y2),"angle":ang})

        # estadísticas angulares
        angs = np.array(angles)
        # normalizar ángulos a [-90,90]
        angs_norm = ((angs + 180) % 180) - 90
        mean_ang = float(np.mean(angs_norm))
        std_ang = float(np.std(angs_norm))

        resumen["estadisticas"]["mean_angle_deg"] = mean_ang
        resumen["estadisticas"]["std_angle_deg"] = std_ang

        # detectar desviaciones significativas (ej. > 2 grados)
        desviadas = []
        umbral_deg = 2.0
        for i,a in enumerate(angs_norm):
            if abs(a - mean_ang) > umbral_deg:
                desviadas.append({"index": i, "angle": float(a), "delta": float(abs(a-mean_ang))})
        resumen["estadisticas"]["num_desviadas"] = len(desviadas)

        # detectar esquinas no rectas: buscar intersecciones de líneas perpendiculares esperadas
        # Simplificación: se aproximará por bounding boxes de contornos rectangulares
        contours, _ = cv2.findContours(bordes, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        rect_count = 0
        rect_problemas = []
        for c in contours:
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.02 * peri, True)
            if len(approx) == 4:
                rect_count += 1
                # calculo ángulos en esquina
                points = [tuple(pt[0]) for pt in approx]
                # calcular 4 ángulos en polígono
                angulos = []
                for i in range(4):
                    p_prev = points[i-1]
                    p = points[i]
                    p_next = points[(i+1)%4]
                    a1 = angle_between(p_prev, p)
                    a2 = angle_between(p, p_next)
                    internal = abs(a2 - a1)
                    # ajustar a [0,180]
                    internal = internal if internal<=180 else 360 - internal
                    angulos.append(internal)
                # si alguna esquina se desvía mucho de 90
                for idx,ia in enumerate(angulos):
                    if abs(ia - 90) > 5:  # umbral 5°
                        rect_problemas.append({"rect_points": points, "corner_index": idx, "angle": ia})
        if rect_count==0:
            resumen["problemas"].append("No se detectaron rectángulos definidos (p.e. habitaciones).")
        else:
            resumen["estadisticas"]["rect_count"] = rect_count
            if rect_problemas:
                resumen["problemas"].append(f"{len(rect_problemas)} esquinas con ángulos no rectos (>5°).")

    # Generar imagen corregida: copia y dibujar
    vis = img.copy()
    # dibujar todas las líneas detectadas en azul
    if lines is not None:
        for l in lines:
            x1,y1,x2,y2 = l[0]
            cv2.line(vis,(x1,y1),(x2,y2),(255,0,0),1)

    # marcar desviadas en rojo (líneas con delta > umbral)
    if lines is not None and len(resumen.get("lineas", []))>0:
        for d in resumen["estadisticas"].get("num_desviadas", 0) and desviadas or []:
            idx = d["index"]
            ln = resumen["lineas"][idx]
            cv2.line(vis,(ln["x1"],ln["y1"]),(ln["x2"],ln["y2"]),(0,0,255),2)  # rojo

    # resaltar rectángulos problemáticos
    for rp in rect_problemas:
        pts = np.array(rp["rect_points"], np.int32)
        cv2.polylines(vis, [pts], isClosed=True, color=(0,0,255), thickness=2)

    # guardar corregida con nombre único
    os.makedirs(salida_dir, exist_ok=True)
    salida = os.path.join(salida_dir, f"{basename}_corr.png")
    cv2.imwrite(salida, vis)

    # construir resumen textual breve (se usará para OpenAI)
    texto_resumen = []
    texto_resumen.append(f"Se detectaron {resumen.get('num_lineas',0)} líneas.")
    stats = resumen.get("estadisticas", {})
    if stats:
        texto_resumen.append(f"Ángulo promedio: {stats.get('mean_angle_deg',0):.2f}°, desviación estándar: {stats.get('std_angle_deg',0):.2f}°.")
        texto_resumen.append(f"Líneas con desviación mayor a {umbral_deg}°: {stats.get('num_desviadas',0)}.")
        if "rect_count" in stats:
            texto_resumen.append(f"Se detectaron {stats['rect_count']} rectángulos (posibles habitaciones/elementos).")
    if resumen.get("problemas"):
        texto_resumen.append("Problemas detectados:")
        for p in resumen["problemas"]:
            texto_resumen.append(f"- {p}")

    resumen["texto_bruto"] = "\n".join(texto_resumen)
    resumen["ruta_corregida"] = salida
    return resumen
