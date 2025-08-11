#generador de cartas gantt

# gantt_generator.py
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment

def crear_gantt_excel(nombre_carpeta="MisExcels", nombre_archivo="gantt.xlsx"):
    os.makedirs(nombre_carpeta, exist_ok=True)
    ruta_archivo = os.path.join(nombre_carpeta, nombre_archivo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Carta Gantt"

    # Encabezados
    dias = [f"Día {i+1}" for i in range(15)]
    ws.append(["Tarea"] + dias)

    # Tareas con inicio y duración
    tareas = [
        ("Planificación", 0, 3),
        ("Diseño", 3, 5),
        ("Desarrollo", 8, 5),
        ("Pruebas", 13, 2)
    ]

    azul = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for tarea, inicio, duracion in tareas:
        fila = [tarea] + ["" for _ in range(15)]
        for i in range(inicio + 1, inicio + duracion + 1):
            fila[i] = "■"
        ws.append(fila)

    # Aplicar color
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=16):
        for cell in row:
            if cell.value == "■":
                cell.fill = azul
                cell.alignment = Alignment(horizontal="center")

    wb.save(ruta_archivo)
    return f"Carta Gantt creada en: {ruta_archivo}"
